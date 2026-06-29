"""
Service buat parsing file Excel Master Customer, Supplier, dan Pekerjaan/BOM
dari format kantor.

Header-nya gak di baris pertama (ada beberapa baris catatan duluan),
jadi kita cari baris header-nya otomatis dengan cocokin teks kolom,
bukan hardcode nomor baris — biar tetap jalan kalau jumlah baris
catatan di atas berubah dikit di file Excel versi lain.
"""

import re
import zipfile
from io import BytesIO
import openpyxl

CUSTOMER_COLUMN_ALIASES = {
    "nama_cust": ["nama cust", "nama customer"],
    "kode_cust": ["kode cust", "kode customer"],
    "contact_person_1": ["contact"],
    "no_telp_1": ["hp/tlp", "hp / tlp", "telp", "tlp"],
    "alamat": ["alamat"],
}

SUPPLIER_COLUMN_ALIASES = {
    "nama_sup": ["nama supl", "nama suplier", "nama supplier"],
    "tipe": ["impor/lokal", "import/lokal"],
    "kode_sup": ["kode supl", "kode suplier", "kode supplier"],
    "direktur": ["direktur"],
    "contact_person_1": ["contact"],
    "no_telp_1": ["tlp/hp", "hp/tlp", "telp/hp"],
    "alamat": ["alamat"],
    "material_dapat_disupply": ["material"],
    "lead_time_produksi": ["lead time produksi (hari)", "lead time produksi"],
    "lead_time_delivery": ["lead time delivery (hari)", "lead time delivery"],
}

MATERIAL_COLUMN_ALIASES = {
    "nama_supplier": ["nama supplier"],
    "kode_sup": ["supplier code"],
    "tipe": ["type"],
    "partnumber": ["partnumber", "part number"],
}


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _extract_int(value):
    """Ambil angka dari value yang kadang ditulis '10 hari', kadang murni angka."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else None


def _normalize_tipe(value) -> str:
    text = _normalize(value)
    if "impor" in text or "import" in text:
        return "import"
    return "lokal"


def _strip_autofilter(file_bytes: bytes) -> bytes:
    """
    Buka file .xlsx sebagai ZIP, hapus elemen <autoFilter> dari tiap sheet XML.
    Beberapa file Excel lama punya setting AutoFilter yang nilainya gak
    sesuai aturan ketat openpyxl, bikin gagal dibaca total walau isinya
    sendiri valid. Karena kita cuma butuh ISI sel, bukan fitur filter-nya,
    elemen itu dibuang aja sebelum diserahin ke openpyxl.
    """
    try:
        input_buffer = BytesIO(file_bytes)
        output_buffer = BytesIO()

        with zipfile.ZipFile(input_buffer, "r") as zin:
            with zipfile.ZipFile(output_buffer, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                        text = data.decode("utf-8", errors="ignore")
                        text = re.sub(r"<autoFilter\b[^>]*?(/>|>.*?</autoFilter>)", "", text, flags=re.DOTALL)
                        data = text.encode("utf-8")
                    zout.writestr(item, data)

        return output_buffer.getvalue()
    except Exception:
        return file_bytes


def _find_header_row(sheet, required_fields: list[str], column_aliases: dict, max_scan_rows: int = 20):
    """
    Cari baris header dengan scan baris-baris awal, cari yang punya
    semua `required_fields` ketemu alias-nya. Balikin (header_row_idx, col_map).

    Pakai index posisi (enumerate) buat nomor kolom, BUKAN cell.column —
    di mode read_only, sel yang gak pernah ditulis di file aslinya kadang
    jadi objek EmptyCell yang gak punya atribut .column/.row sama sekali.
    """
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan_rows), start=1):
        found: dict[str, int] = {}
        for col_idx, cell in enumerate(row, start=1):
            text = _normalize(cell.value)
            if not text:
                continue
            for field, aliases in column_aliases.items():
                if any(alias in text for alias in aliases):
                    found[field] = col_idx

        if all(field in found for field in required_fields):
            return row_idx, found

    return None, {}


def parse_customer_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse file Excel Master Customer, balikin list of dict siap insert.
    Baris yang nama_cust atau kode_cust-nya kosong otomatis di-skip
    (ini termasuk baris 'reserved' yang cuma ada kode doang).
    """
    file_bytes = _strip_autofilter(file_bytes)
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook.worksheets[0]

    header_row_idx, col_map = _find_header_row(sheet, ["kode_cust", "nama_cust"], CUSTOMER_COLUMN_ALIASES)

    if header_row_idx is None:
        raise ValueError(
            "Header kolom 'Nama Cust' dan 'Kode Cust' tidak ditemukan di 20 baris pertama file ini"
        )

    def get_value(row_cells, field: str):
        col = col_map.get(field)
        if not col:
            return None
        value = row_cells[col - 1].value
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    results = []
    for row_cells in sheet.iter_rows(min_row=header_row_idx + 1):
        nama_cust = get_value(row_cells, "nama_cust")
        kode_cust = get_value(row_cells, "kode_cust")

        if not nama_cust or not kode_cust:
            continue

        results.append({
            "nama_cust": nama_cust,
            "kode_cust": kode_cust,
            "contact_person_1": get_value(row_cells, "contact_person_1"),
            "no_telp_1": get_value(row_cells, "no_telp_1"),
            "alamat": get_value(row_cells, "alamat"),
        })

    return results


def parse_supplier_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse file Excel Master Supplier, balikin list of dict siap insert.
    Catatan: kolom 'kategori' WAJIB diisi di database kita, tapi gak ada
    di format Excel ini — di-default jadi 'lainnya', bisa dikoreksi manual
    lewat form edit nanti.
    """
    file_bytes = _strip_autofilter(file_bytes)
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook.worksheets[0]

    header_row_idx, col_map = _find_header_row(sheet, ["kode_sup", "nama_sup"], SUPPLIER_COLUMN_ALIASES)

    if header_row_idx is None:
        raise ValueError(
            "Header kolom 'Nama Supl' dan 'Kode Supl' tidak ditemukan di 20 baris pertama file ini"
        )

    def get_value(row_cells, field: str):
        col = col_map.get(field)
        if not col:
            return None
        value = row_cells[col - 1].value
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    results = []
    for row_cells in sheet.iter_rows(min_row=header_row_idx + 1):
        nama_sup = get_value(row_cells, "nama_sup")
        kode_sup = get_value(row_cells, "kode_sup")

        if not nama_sup or not kode_sup:
            continue

        results.append({
            "nama_sup": nama_sup,
            "kode_sup": kode_sup,
            "tipe": _normalize_tipe(get_value(row_cells, "tipe")),
            "direktur": get_value(row_cells, "direktur"),
            "contact_person_1": get_value(row_cells, "contact_person_1"),
            "no_telp_1": get_value(row_cells, "no_telp_1"),
            "alamat": get_value(row_cells, "alamat"),
            "material_dapat_disupply": get_value(row_cells, "material_dapat_disupply"),
            "lead_time_produksi": _extract_int(get_value(row_cells, "lead_time_produksi")),
            "lead_time_delivery": _extract_int(get_value(row_cells, "lead_time_delivery")),
        })

    return results


def parse_material_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse sheet 'master part number per supp'. Sheet ini punya ratusan
    kolom tambahan di sebelah kanan (sampai kolom FS) yang isinya cuma
    tabel bantu buat formula Excel asli (dropdown, agregasi per supplier)
    — BUKAN sumber data. Sumber aslinya cuma 4 kolom pertama: nama
    supplier, supplier code, type, partnumber.

    Field 'satuan_terkecil' gak ada di Excel ini, di-default:
    type 'kabel' -> 'm', selain itu (housing/terminal/lainlain) -> 'pcs'.
    """
    file_bytes = _strip_autofilter(file_bytes)
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)

    if "master part number per supp" not in workbook.sheetnames:
        raise ValueError("Sheet 'master part number per supp' tidak ditemukan di file ini")

    sheet = workbook["master part number per supp"]

    header_row_idx, col_map = _find_header_row(
        sheet, ["nama_supplier", "kode_sup", "partnumber"], MATERIAL_COLUMN_ALIASES
    )

    if header_row_idx is None:
        raise ValueError(
            "Header 'nama supplier', 'supplier code', dan 'partnumber' "
            "tidak ditemukan di 20 baris pertama sheet ini"
        )

    def get_value(row_cells, field: str):
        col = col_map.get(field)
        if not col:
            return None
        value = row_cells[col - 1].value
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    results = []
    for row_cells in sheet.iter_rows(min_row=header_row_idx + 1):
        kode_sup = get_value(row_cells, "kode_sup")
        partnumber = get_value(row_cells, "partnumber")
        tipe = get_value(row_cells, "tipe")

        if not kode_sup or not partnumber:
            continue

        satuan_terkecil = "m" if tipe and tipe.lower() == "kabel" else "pcs"
        kategori_map = {"kabel": "kabel", "housing": "housing", "terminal": "terminal", "lainlain": "lainnya"}
        kategori = kategori_map.get(tipe.lower(), "lainnya") if tipe else "lainnya"

        results.append({
            "nama_material": partnumber,
            "kode_sup": kode_sup,
            "satuan_terkecil": satuan_terkecil,
            "kategori": kategori,
        })

    return results


BOM_SHEET_NAMES = [
    "satuan BOM kabel",
    "satuan BOM housing ",  # ada trailing space di nama sheet aslinya
    "satuan BOM terminal",
    "satuan BOM lainnya",
]


def _parse_bom_sheet(sheet) -> list[dict]:
    """
    Parse 1 sheet BOM (kabel/housing/terminal/lainnya) — semuanya pakai
    layout yang sama: nama material di baris TEPAT SEBELUM baris yang punya
    label 'p/n cust' (kolom A) + 'cust code' (kolom B). Data mulai 1 baris
    setelah itu. Ada 1 kolom bantu terpisah jauh di sebelah kanan (kolom
    label kayak 'material kabel') yang isinya list vertikal gak terkait —
    otomatis ke-skip karena baris header materialnya kosong di kolom itu.
    """
    anchor_row_idx = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
        if len(row) < 2:
            continue
        col_a = _normalize(row[0].value)
        col_b = _normalize(row[1].value)
        if col_a == "p/n cust" and "cust code" in col_b:
            anchor_row_idx = row_idx
            break

    if anchor_row_idx is None:
        return []

    header_row_idx = anchor_row_idx - 1
    material_columns: dict[int, str] = {}
    header_row = next(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx))
    for col_idx, cell in enumerate(header_row, start=1):
        if col_idx <= 3:
            continue
        if cell.value:
            material_columns[col_idx] = str(cell.value).strip()

    results = []
    for row_cells in sheet.iter_rows(min_row=anchor_row_idx + 1):
        kode_barang = row_cells[0].value
        kode_cust = row_cells[1].value

        if not kode_barang or not kode_cust:
            continue

        for col_idx, nama_material in material_columns.items():
            if col_idx > len(row_cells):
                continue
            cell = row_cells[col_idx - 1]
            if cell.value is not None:
                results.append({
                    "kode_barang": str(kode_barang).strip(),
                    "kode_cust": str(kode_cust).strip(),
                    "nama_material": nama_material,
                    "penggunaan": cell.value,
                })

    return results


def parse_bom_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse 4 sheet BOM (kabel/housing/terminal/lainnya) dari file Master
    Pekerjaan dan BOM, gabungin jadi 1 list flat.

    Balikin: [{"kode_barang": "...", "kode_cust": "...",
               "nama_material": "...", "penggunaan": ...}, ...]
    """
    file_bytes = _strip_autofilter(file_bytes)
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)

    results = []
    for sheet_name in BOM_SHEET_NAMES:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        results.extend(_parse_bom_sheet(sheet))

    return results


def parse_pekerjaan_excel(file_bytes: bytes) -> dict:
    """
    Parse sheet 'input pekerjaan' dari file Master Pekerjaan dan BOM.
    Strukturnya beda dari Customer/Supplier: nama pekerjaan ada di baris
    TEPAT SEBELUM baris yang punya label 'P/N CUST' (kolom A) dan
    'kode cust' (kolom B). Data mulai 1 baris setelah baris label itu.

    Balikin dict: {"rows": [...], "job_order": [...]}
    - rows: [{"kode_barang": "...", "kode_cust": "...",
              "jasa_entries": [{"jenis_pekerjaan": "...", "urutan": ...}, ...]}]
    - job_order: list nama pekerjaan urut sesuai posisi kolom kiri-ke-kanan
      di Excel (dipakai buat ngisi urutan pas Jasa baru di-create otomatis)
    """
    file_bytes = _strip_autofilter(file_bytes)
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)

    if "input pekerjaan" not in workbook.sheetnames:
        raise ValueError("Sheet 'input pekerjaan' tidak ditemukan di file ini")

    sheet = workbook["input pekerjaan"]

    anchor_row_idx = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
        if len(row) < 2:
            continue
        col_a = _normalize(row[0].value)
        col_b = _normalize(row[1].value)
        if col_a == "p/n cust" and "kode cust" in col_b:
            anchor_row_idx = row_idx
            break

    if anchor_row_idx is None:
        raise ValueError(
            "Header 'P/N CUST' dan 'kode cust' tidak ditemukan di 20 baris pertama sheet 'input pekerjaan'"
        )

    job_header_row_idx = anchor_row_idx - 1
    job_columns: dict[int, str] = {}
    job_header_row = next(sheet.iter_rows(min_row=job_header_row_idx, max_row=job_header_row_idx))
    for col_idx, cell in enumerate(job_header_row, start=1):
        if col_idx <= 2:
            continue
        if cell.value:
            job_columns[col_idx] = str(cell.value).strip()

    results = []
    for row_cells in sheet.iter_rows(min_row=anchor_row_idx + 1):
        kode_barang = row_cells[0].value
        kode_cust = row_cells[1].value

        if not kode_barang or not kode_cust:
            continue

        jasa_entries = []
        for col_idx, job_name in job_columns.items():
            cell = row_cells[col_idx - 1]
            if cell.value is not None:
                jasa_entries.append({"jenis_pekerjaan": job_name, "urutan": cell.value})

        results.append({
            "kode_barang": str(kode_barang).strip(),
            "kode_cust": str(kode_cust).strip(),
            "jasa_entries": jasa_entries,
        })

    return {"rows": results, "job_order": list(job_columns.values())}
